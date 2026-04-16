#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Générateur de Quittances de Loyer et Avis d'Échéance
=====================================================
Lit les données depuis Gestion_Locative.xlsx et génère des PDF.

Usage:
    python generer_documents.py quittance --bail BAIL-001 --periode "Janvier 2026"
    python generer_documents.py avis --bail BAIL-001 --periode "Février 2026"
    python generer_documents.py quittance --all --periode "Janvier 2026"
    python generer_documents.py avis --all --periode "Mars 2026"
"""

import argparse
import os
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm, cm
from reportlab.lib.colors import HexColor
from reportlab.pdfgen import canvas
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT, TA_JUSTIFY
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Frame, PageTemplate
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# ── Configuration ──────────────────────────────────────────────
EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Gestion_Locative.xlsx")
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Documents_PDF")

PRIMARY_COLOR = HexColor('#2C3E50')
ACCENT_COLOR = HexColor('#3498DB')
LIGHT_BG = HexColor('#F8F9FA')
BORDER_COLOR = HexColor('#BDC3C7')

MOIS_FR = {
    "janvier": 1, "février": 2, "mars": 3, "avril": 4,
    "mai": 5, "juin": 6, "juillet": 7, "août": 8,
    "septembre": 9, "octobre": 10, "novembre": 11, "décembre": 12
}

def format_euro(amount):
    if amount is None:
        return "0,00 €"
    return f"{amount:,.2f} €".replace(",", " ").replace(".", ",")

def format_date_fr(d):
    if isinstance(d, datetime):
        d = d.date()
    if isinstance(d, date):
        mois = list(MOIS_FR.keys())[d.month - 1]
        return f"{d.day} {mois} {d.year}"
    return str(d)

def nombre_en_lettres(n):
    """Conversion simple d'un nombre en lettres (pour les montants)."""
    unites = ["", "un", "deux", "trois", "quatre", "cinq", "six", "sept", "huit", "neuf",
              "dix", "onze", "douze", "treize", "quatorze", "quinze", "seize",
              "dix-sept", "dix-huit", "dix-neuf"]
    dizaines = ["", "dix", "vingt", "trente", "quarante", "cinquante",
                "soixante", "soixante", "quatre-vingt", "quatre-vingt"]

    if n == 0:
        return "zéro"

    def _convert(num):
        if num < 20:
            return unites[num]
        elif num < 70:
            d, u = divmod(num, 10)
            sep = "-" if u else ""
            return dizaines[d] + sep + unites[u]
        elif num < 80:
            u = num - 60
            sep = " et " if u == 1 else "-"
            if u == 0:
                return "soixante-dix"
            return "soixante" + sep + unites[u + 10] if u < 10 else "soixante-" + unites[u]
        elif num < 100:
            u = num - 80
            if u == 0:
                return "quatre-vingts"
            return "quatre-vingt-" + (_convert(u) if u >= 10 else unites[u])
        elif num < 1000:
            c, r = divmod(num, 100)
            prefix = "" if c == 1 else unites[c] + " "
            cents = "cent" if r == 0 and c > 1 else "cent"
            if c == 1:
                prefix = ""
                cents = "cent"
            rest = " " + _convert(r) if r else ("s" if c > 1 else "")
            if r:
                return (prefix + cents + " " + _convert(r)).strip()
            else:
                return (prefix + cents + ("s" if c > 1 else "")).strip()
        elif num < 10000:
            m, r = divmod(num, 1000)
            prefix = "mille" if m == 1 else unites[m] + " mille"
            rest = " " + _convert(r) if r else ""
            return prefix + rest
        return str(num)

    euros = int(n)
    centimes = round((n - euros) * 100)
    result = _convert(euros) + " euro" + ("s" if euros > 1 else "")
    if centimes > 0:
        result += " et " + _convert(centimes) + " centime" + ("s" if centimes > 1 else "")
    return result


# ── Chargement des données ─────────────────────────────────────
def load_data():
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    # Propriétaire
    ws = wb["Propriétaire"]
    proprio = {
        "nom": ws.cell(3, 2).value or "",
        "adresse": ws.cell(4, 2).value or "",
        "cp": ws.cell(5, 2).value or "",
        "ville": ws.cell(6, 2).value or "",
        "tel": ws.cell(7, 2).value or "",
        "email": ws.cell(8, 2).value or "",
        "siret": ws.cell(9, 2).value or "",
    }

    # Baux
    ws = wb["Baux"]
    baux = {}
    for row in range(3, ws.max_row + 1):
        bail_id = ws.cell(row, 1).value
        if not bail_id:
            continue
        baux[bail_id] = {
            "id": bail_id,
            "bien_id": ws.cell(row, 2).value,
            "bien_nom": ws.cell(row, 3).value,
            "loc_id": ws.cell(row, 4).value,
            "loc_nom": ws.cell(row, 5).value,
            "date_debut": ws.cell(row, 6).value,
            "date_fin": ws.cell(row, 7).value,
            "loyer_hc": ws.cell(row, 8).value or 0,
            "charges": ws.cell(row, 9).value or 0,
            "total": (ws.cell(row, 8).value or 0) + (ws.cell(row, 9).value or 0),
            "jour_echeance": ws.cell(row, 11).value or 1,
            "depot_garantie": ws.cell(row, 12).value or 0,
            "statut": ws.cell(row, 13).value,
        }

    # Locataires
    ws = wb["Locataires"]
    locataires = {}
    for row in range(3, ws.max_row + 1):
        loc_id = ws.cell(row, 1).value
        if not loc_id:
            continue
        locataires[loc_id] = {
            "id": loc_id,
            "civilite": ws.cell(row, 2).value or "",
            "nom": ws.cell(row, 3).value or "",
            "prenom": ws.cell(row, 4).value or "",
            "adresse": ws.cell(row, 5).value or "",
            "cp": ws.cell(row, 6).value or "",
            "ville": ws.cell(row, 7).value or "",
        }

    # Biens
    ws = wb["Biens"]
    biens = {}
    for row in range(3, ws.max_row + 1):
        bien_id = ws.cell(row, 1).value
        if not bien_id:
            continue
        biens[bien_id] = {
            "id": bien_id,
            "designation": ws.cell(row, 2).value or "",
            "type": ws.cell(row, 3).value or "",
            "adresse": ws.cell(row, 4).value or "",
            "cp": ws.cell(row, 5).value or "",
            "ville": ws.cell(row, 6).value or "",
            "surface": ws.cell(row, 7).value or "",
        }

    wb.close()
    return proprio, baux, locataires, biens


# ── Génération Quittance ───────────────────────────────────────
def generer_quittance(bail_id, periode, proprio, baux, locataires, biens):
    if bail_id not in baux:
        print(f"Erreur: bail {bail_id} non trouvé.")
        return None

    bail = baux[bail_id]
    loc = locataires.get(bail["loc_id"], {})
    bien = biens.get(bail["bien_id"], {})

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filename = f"Quittance_{bail_id}_{periode.replace(' ', '_')}.pdf"
    filepath = os.path.join(OUTPUT_DIR, filename)

    c = canvas.Canvas(filepath, pagesize=A4)
    w, h = A4

    # ── En-tête ──
    c.setFillColor(PRIMARY_COLOR)
    c.rect(0, h - 80, w, 80, fill=True, stroke=False)
    c.setFillColor(HexColor('#FFFFFF'))
    c.setFont("Helvetica-Bold", 22)
    c.drawCentredString(w / 2, h - 45, "QUITTANCE DE LOYER")
    c.setFont("Helvetica", 11)
    c.drawCentredString(w / 2, h - 65, f"Période : {periode}")

    y = h - 110

    # ── Propriétaire ──
    c.setFillColor(ACCENT_COLOR)
    c.rect(30, y - 5, 250, 20, fill=True, stroke=False)
    c.setFillColor(HexColor('#FFFFFF'))
    c.setFont("Helvetica-Bold", 10)
    c.drawString(35, y, "BAILLEUR")
    y -= 25
    c.setFillColor(PRIMARY_COLOR)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(35, y, proprio["nom"] if proprio["nom"] else "[ À compléter ]")
    y -= 15
    c.setFont("Helvetica", 9)
    if proprio["adresse"]:
        c.drawString(35, y, proprio["adresse"])
        y -= 13
    if proprio["cp"] or proprio["ville"]:
        c.drawString(35, y, f"{proprio['cp']} {proprio['ville']}")
        y -= 13
    if proprio["siret"]:
        c.drawString(35, y, f"SIRET : {proprio['siret']}")
        y -= 13

    # ── Locataire ──
    y_loc = h - 110
    c.setFillColor(ACCENT_COLOR)
    c.rect(310, y_loc - 5, 250, 20, fill=True, stroke=False)
    c.setFillColor(HexColor('#FFFFFF'))
    c.setFont("Helvetica-Bold", 10)
    c.drawString(315, y_loc, "LOCATAIRE")
    y_loc -= 25
    c.setFillColor(PRIMARY_COLOR)
    c.setFont("Helvetica-Bold", 10)
    loc_fullname = f"{loc.get('civilite', '')} {loc.get('prenom', '')} {loc.get('nom', '')}".strip()
    c.drawString(315, y_loc, loc_fullname or bail["loc_nom"])
    y_loc -= 15
    c.setFont("Helvetica", 9)
    if loc.get("adresse"):
        c.drawString(315, y_loc, loc["adresse"])
        y_loc -= 13
    if loc.get("cp") or loc.get("ville"):
        c.drawString(315, y_loc, f"{loc.get('cp', '')} {loc.get('ville', '')}")

    # ── Ligne de séparation ──
    y = min(y, y_loc) - 25
    c.setStrokeColor(BORDER_COLOR)
    c.line(30, y, w - 30, y)
    y -= 25

    # ── Bien ──
    c.setFillColor(LIGHT_BG)
    c.rect(30, y - 60, w - 60, 75, fill=True, stroke=False)
    c.setStrokeColor(BORDER_COLOR)
    c.rect(30, y - 60, w - 60, 75, fill=False, stroke=True)

    c.setFillColor(PRIMARY_COLOR)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(40, y + 2, "Bien loué :")
    c.setFont("Helvetica", 10)
    c.drawString(120, y + 2, f"{bien.get('designation', bail['bien_nom'])} ({bien.get('type', '')})")
    y -= 15
    c.drawString(40, y + 2, f"Adresse : {bien.get('adresse', '')} - {bien.get('cp', '')} {bien.get('ville', '')}")
    y -= 15
    if bien.get('surface'):
        c.drawString(40, y + 2, f"Surface : {bien['surface']} m²")
    y -= 15
    c.drawString(40, y + 2, f"Bail n° : {bail_id}")

    y -= 40

    # ── Détail du paiement ──
    c.setFillColor(ACCENT_COLOR)
    c.rect(30, y - 5, w - 60, 22, fill=True, stroke=False)
    c.setFillColor(HexColor('#FFFFFF'))
    c.setFont("Helvetica-Bold", 11)
    c.drawCentredString(w / 2, y, "DÉTAIL DU PAIEMENT")
    y -= 35

    # Table
    loyer = bail["loyer_hc"]
    charges = bail["charges"]
    total = loyer + charges

    data = [
        ["Désignation", "Montant"],
        ["Loyer hors charges", format_euro(loyer)],
        ["Provision pour charges", format_euro(charges)],
        ["", ""],
        ["TOTAL", format_euro(total)],
    ]

    table = Table(data, colWidths=[350, 150])
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('BACKGROUND', (0, 0), (-1, 0), ACCENT_COLOR),
        ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#FFFFFF')),
        ('BACKGROUND', (0, -1), (-1, -1), HexColor('#EBF5FB')),
        ('LINEBELOW', (0, 0), (-1, 0), 1, BORDER_COLOR),
        ('LINEABOVE', (0, -1), (-1, -1), 1.5, PRIMARY_COLOR),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
    ]))

    tw, th = table.wrap(0, 0)
    table.drawOn(c, 50, y - th)
    y = y - th - 20

    # Montant en lettres
    c.setFont("Helvetica-Oblique", 9)
    c.setFillColor(PRIMARY_COLOR)
    c.drawString(40, y, f"Soit la somme de : {nombre_en_lettres(total)}")
    y -= 35

    # ── Texte légal ──
    c.setStrokeColor(BORDER_COLOR)
    c.line(30, y, w - 30, y)
    y -= 20
    c.setFont("Helvetica", 8)
    c.setFillColor(HexColor('#666666'))
    text_legal = (
        f"Le bailleur, {proprio['nom'] or '_______________'}, propriétaire du logement désigné ci-dessus, "
        f"déclare avoir reçu de {loc_fullname or bail['loc_nom']} la somme de {format_euro(total)} "
        f"au titre du paiement du loyer et des charges pour la période de {periode}."
    )
    # Word wrap
    from reportlab.lib.utils import simpleSplit
    lines = simpleSplit(text_legal, "Helvetica", 8, w - 80)
    for line in lines:
        c.drawString(40, y, line)
        y -= 12

    y -= 10
    c.drawString(40, y, "Cette quittance annule tous les reçus qui auraient pu être délivrés pour acomptes versés au titre de cette même période.")
    y -= 12
    c.drawString(40, y, "Elle est délivrée sous réserve de tous droits, actions et procédures du bailleur.")

    y -= 40

    # ── Date et signature ──
    today = datetime.now()
    c.setFillColor(PRIMARY_COLOR)
    c.setFont("Helvetica", 10)
    c.drawString(350, y, f"Fait à {proprio['ville'] or '___________'}")
    y -= 15
    c.drawString(350, y, f"Le {format_date_fr(today.date())}")
    y -= 30
    c.drawString(350, y, "Signature du bailleur :")
    y -= 40
    c.setStrokeColor(BORDER_COLOR)
    c.setDash(3, 3)
    c.line(350, y, 530, y)

    # ── Pied de page ──
    c.setDash()
    c.setFillColor(PRIMARY_COLOR)
    c.rect(0, 0, w, 30, fill=True, stroke=False)
    c.setFillColor(HexColor('#FFFFFF'))
    c.setFont("Helvetica", 7)
    c.drawCentredString(w / 2, 12, "Document généré automatiquement - Gestion Locative")

    c.save()
    print(f"Quittance générée : {filepath}")
    return filepath


# ── Génération Avis d'Échéance ─────────────────────────────────
def generer_avis(bail_id, periode, proprio, baux, locataires, biens):
    if bail_id not in baux:
        print(f"Erreur: bail {bail_id} non trouvé.")
        return None

    bail = baux[bail_id]
    loc = locataires.get(bail["loc_id"], {})
    bien = biens.get(bail["bien_id"], {})

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filename = f"Avis_Echeance_{bail_id}_{periode.replace(' ', '_')}.pdf"
    filepath = os.path.join(OUTPUT_DIR, filename)

    c = canvas.Canvas(filepath, pagesize=A4)
    w, h = A4

    # ── En-tête ──
    c.setFillColor(HexColor('#E74C3C'))
    c.rect(0, h - 80, w, 80, fill=True, stroke=False)
    c.setFillColor(HexColor('#FFFFFF'))
    c.setFont("Helvetica-Bold", 22)
    c.drawCentredString(w / 2, h - 45, "AVIS D'ÉCHÉANCE")
    c.setFont("Helvetica", 11)
    c.drawCentredString(w / 2, h - 65, f"Période : {periode}")

    y = h - 110

    # ── Propriétaire ──
    c.setFillColor(PRIMARY_COLOR)
    c.rect(30, y - 5, 250, 20, fill=True, stroke=False)
    c.setFillColor(HexColor('#FFFFFF'))
    c.setFont("Helvetica-Bold", 10)
    c.drawString(35, y, "BAILLEUR")
    y -= 25
    c.setFillColor(PRIMARY_COLOR)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(35, y, proprio["nom"] if proprio["nom"] else "[ À compléter ]")
    y -= 15
    c.setFont("Helvetica", 9)
    if proprio["adresse"]:
        c.drawString(35, y, proprio["adresse"])
        y -= 13
    if proprio["cp"] or proprio["ville"]:
        c.drawString(35, y, f"{proprio['cp']} {proprio['ville']}")

    # ── Locataire ──
    y_loc = h - 110
    c.setFillColor(PRIMARY_COLOR)
    c.rect(310, y_loc - 5, 250, 20, fill=True, stroke=False)
    c.setFillColor(HexColor('#FFFFFF'))
    c.setFont("Helvetica-Bold", 10)
    c.drawString(315, y_loc, "LOCATAIRE")
    y_loc -= 25
    c.setFillColor(PRIMARY_COLOR)
    c.setFont("Helvetica-Bold", 10)
    loc_fullname = f"{loc.get('civilite', '')} {loc.get('prenom', '')} {loc.get('nom', '')}".strip()
    c.drawString(315, y_loc, loc_fullname or bail["loc_nom"])
    y_loc -= 15
    c.setFont("Helvetica", 9)
    if loc.get("adresse"):
        c.drawString(315, y_loc, loc["adresse"])
        y_loc -= 13
    if loc.get("cp") or loc.get("ville"):
        c.drawString(315, y_loc, f"{loc.get('cp', '')} {loc.get('ville', '')}")

    y = min(y, y_loc) - 25
    c.setStrokeColor(BORDER_COLOR)
    c.line(30, y, w - 30, y)
    y -= 25

    # ── Bien ──
    c.setFillColor(LIGHT_BG)
    c.rect(30, y - 45, w - 60, 60, fill=True, stroke=False)
    c.setStrokeColor(BORDER_COLOR)
    c.rect(30, y - 45, w - 60, 60, fill=False, stroke=True)

    c.setFillColor(PRIMARY_COLOR)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(40, y + 2, "Bien loué :")
    c.setFont("Helvetica", 10)
    c.drawString(120, y + 2, f"{bien.get('designation', bail['bien_nom'])} ({bien.get('type', '')})")
    y -= 15
    c.drawString(40, y + 2, f"Adresse : {bien.get('adresse', '')} - {bien.get('cp', '')} {bien.get('ville', '')}")
    y -= 15
    c.drawString(40, y + 2, f"Bail n° : {bail_id}")

    y -= 40

    # ── Info d'échéance ──
    # Parse period to get due date
    parts = periode.strip().split()
    mois_name = parts[0].lower() if parts else ""
    annee = int(parts[1]) if len(parts) > 1 else datetime.now().year
    mois_num = MOIS_FR.get(mois_name, 1)
    jour = bail.get("jour_echeance", 1) or 1

    c.setFillColor(HexColor('#FADBD8'))
    c.rect(30, y - 40, w - 60, 55, fill=True, stroke=False)
    c.setStrokeColor(HexColor('#E74C3C'))
    c.rect(30, y - 40, w - 60, 55, fill=False, stroke=True)

    c.setFillColor(HexColor('#E74C3C'))
    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(w / 2, y, f"Date d'échéance : {jour:02d}/{mois_num:02d}/{annee}")
    y -= 20
    c.setFont("Helvetica-Bold", 11)
    c.setFillColor(PRIMARY_COLOR)
    c.drawCentredString(w / 2, y, f"Montant à régler : {format_euro(bail['total'])}")

    y -= 50

    # ── Détail ──
    c.setFillColor(PRIMARY_COLOR)
    c.rect(30, y - 5, w - 60, 22, fill=True, stroke=False)
    c.setFillColor(HexColor('#FFFFFF'))
    c.setFont("Helvetica-Bold", 11)
    c.drawCentredString(w / 2, y, "DÉTAIL DE L'ÉCHÉANCE")
    y -= 35

    loyer = bail["loyer_hc"]
    charges = bail["charges"]
    total = loyer + charges

    data = [
        ["Désignation", "Montant"],
        ["Loyer hors charges", format_euro(loyer)],
        ["Provision pour charges", format_euro(charges)],
        ["", ""],
        ["TOTAL À RÉGLER", format_euro(total)],
    ]

    table = Table(data, colWidths=[350, 150])
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('BACKGROUND', (0, 0), (-1, 0), PRIMARY_COLOR),
        ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#FFFFFF')),
        ('BACKGROUND', (0, -1), (-1, -1), HexColor('#FADBD8')),
        ('LINEBELOW', (0, 0), (-1, 0), 1, BORDER_COLOR),
        ('LINEABOVE', (0, -1), (-1, -1), 1.5, HexColor('#E74C3C')),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
    ]))

    tw, th = table.wrap(0, 0)
    table.drawOn(c, 50, y - th)
    y = y - th - 20

    c.setFont("Helvetica-Oblique", 9)
    c.setFillColor(PRIMARY_COLOR)
    c.drawString(40, y, f"Soit la somme de : {nombre_en_lettres(total)}")
    y -= 40

    # ── Mentions ──
    c.setStrokeColor(BORDER_COLOR)
    c.line(30, y, w - 30, y)
    y -= 20
    c.setFont("Helvetica", 8)
    c.setFillColor(HexColor('#666666'))

    mentions = [
        "Cet avis d'échéance ne constitue pas une quittance de loyer.",
        "La quittance vous sera délivrée après encaissement effectif du paiement.",
        f"Merci d'effectuer votre règlement avant le {jour:02d}/{mois_num:02d}/{annee}.",
        "",
        "En cas de difficultés de paiement, nous vous invitons à prendre contact avec nous dans les meilleurs délais.",
    ]
    for line in mentions:
        c.drawString(40, y, line)
        y -= 12

    y -= 30

    # ── Date ──
    today = datetime.now()
    c.setFillColor(PRIMARY_COLOR)
    c.setFont("Helvetica", 10)
    c.drawString(350, y, f"Fait à {proprio['ville'] or '___________'}")
    y -= 15
    c.drawString(350, y, f"Le {format_date_fr(today.date())}")

    # ── Pied de page ──
    c.setFillColor(HexColor('#E74C3C'))
    c.rect(0, 0, w, 30, fill=True, stroke=False)
    c.setFillColor(HexColor('#FFFFFF'))
    c.setFont("Helvetica", 7)
    c.drawCentredString(w / 2, 12, "Document généré automatiquement - Gestion Locative")

    c.save()
    print(f"Avis d'échéance généré : {filepath}")
    return filepath


# ── Main ───────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Générateur de quittances et avis d'échéance")
    parser.add_argument("type", choices=["quittance", "avis"], help="Type de document")
    parser.add_argument("--bail", help="ID du bail (ex: BAIL-001)")
    parser.add_argument("--all", action="store_true", help="Générer pour tous les baux actifs")
    parser.add_argument("--periode", required=True, help="Période (ex: 'Janvier 2026')")

    args = parser.parse_args()

    if not args.bail and not args.all:
        print("Erreur: spécifiez --bail BAIL-ID ou --all")
        sys.exit(1)

    proprio, baux, locataires, biens = load_data()

    if args.all:
        bail_ids = [bid for bid, b in baux.items() if b["statut"] == "Actif"]
    else:
        bail_ids = [args.bail]

    gen_func = generer_quittance if args.type == "quittance" else generer_avis

    generated = []
    for bid in bail_ids:
        result = gen_func(bid, args.periode, proprio, baux, locataires, biens)
        if result:
            generated.append(result)

    print(f"\n{'='*50}")
    print(f"{len(generated)} document(s) généré(s) dans : {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
